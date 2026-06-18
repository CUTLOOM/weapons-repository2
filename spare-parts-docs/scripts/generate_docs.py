import sys, openpyxl, os, io, re, argparse
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from collections import OrderedDict
sys.stdout.reconfigure(encoding="utf-8")

DEFAULT_MASTER = r"D:\1. 德源\1. 订单\德源-GIGA\6. 产品信息汇总\LUTZ补件唛头汇总2026-6-9 更新.xlsx"

def add_image(ws, img_data, row_1b, col_char, col_w_chars, row_h_pt):
    ni = XLImage(io.BytesIO(img_data))
    cw_px = col_w_chars * 7
    ch_px = row_h_pt * 96 / 72
    cw_e = pixels_to_EMU(int(cw_px))
    ch_e = pixels_to_EMU(int(ch_px))
    ratio = min(cw_px*0.7/ni.width, ch_px*0.7/ni.height)
    dw = int(ni.width*ratio)
    dh = int(ni.height*ratio)
    ni.width = dw
    ni.height = dh
    dw_e = pixels_to_EMU(dw)
    dh_e = pixels_to_EMU(dh)
    ox_e = max(0, (cw_e-dw_e)//2)
    oy_e = max(0, (ch_e-dh_e)//2)
    ci = ord(col_char.upper()) - ord("A")
    anc = OneCellAnchor()
    anc._from = AnchorMarker(col=ci, row=row_1b-1, colOff=ox_e, rowOff=oy_e)
    anc.ext = XDRPositiveSize2D(cx=dw_e, cy=dh_e)
    ni.anchor = anc
    ws._images.append(ni)

def generate(pi_path, master_path, order_no=None):
    wb_pi = openpyxl.load_workbook(pi_path)
    ws_pi = wb_pi.active
    groups = OrderedDict()
    for r in range(12, ws_pi.max_row + 1):
        item_no = ws_pi.cell(r, 2).value
        design = ws_pi.cell(r, 3).value
        qty = ws_pi.cell(r, 6).value
        if not item_no or not qty:
            continue
        dn = str(design).split(chr(10))[0].strip()
        if dn not in groups:
            groups[dn] = {"item_nos": [], "total_qty": 0}
        groups[dn]["item_nos"].append(str(item_no).strip())
        groups[dn]["total_qty"] += int(qty)

    if not order_no:
        order_no = str(ws_pi.cell(8, 13).value or "UNKNOWN").strip()

    set_counts = {k: max(1, round(g["total_qty"]*0.03)) for k, g in groups.items()}
    article_strs = {}
    for k, g in groups.items():
        base = list(OrderedDict.fromkeys([p.split("/")[0] for p in g["item_nos"]]))
        sfx = list(OrderedDict.fromkeys([p.split("/")[1] for p in g["item_nos"] if "/" in p]))
        article_strs[k] = (base[0] + "/" + "/".join(sfx)) if sfx else "/".join(base)

    wb_master = openpyxl.load_workbook(master_path)
    master_data = {}
    for key in groups:
        for sn in wb_master.sheetnames:
            if key.lower() in sn.lower() or any(kw in sn for kw in key.split("-")):
                ws = wb_master[sn]
                parts = []
                for r in range(5, ws.max_row + 1):
                    desc = ws.cell(r, 2).value
                    qt = ws.cell(r, 4).value
                    if not desc:
                        continue
                    ps = 0
                    if qt:
                        m = re.search(r"(\d+)\s*[xX]\s*(\d+)\s*=\s*(\d+)", str(qt))
                        if m:
                            ps = int(m.group(2))
                    parts.append({"desc": str(desc).strip(), "per_set": ps})
                imgs = []
                for img in getattr(ws, "_images", []):
                    try:
                        imgs.append({"from_row": img.anchor._from.row, "data": img._data()})
                    except:
                        pass
                master_data[key] = {"parts": parts, "images": imgs}
                break

    out_dir = os.path.dirname(os.path.abspath(pi_path))
    h_f = Font(name="Arial", size=16, color="FF000000", bold=True)
    h_fb = Font(name="Arial", size=18, color="FF000000", bold=True)
    ch_f = Font(name="Arial", size=14, color="FF000000", bold=True)
    ch_fa = Font(name="Arial", size=16, color="FF000000", bold=True)
    d_f = Font(name="SimSun", size=11, color="FF000000")
    q_f = Font(name="Arial", size=12, color="FF000000")
    a_f = Font(name="SimSun", size=11, color="FF000000")
    ac = Alignment(horizontal="center", vertical="center")
    acw = Alignment(horizontal="center", vertical="center", wrap_text=True)
    al = Alignment(horizontal="left", vertical="center")
    alw = Alignment(horizontal="left", vertical="center", wrap_text=True)
    tb = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))
    yf = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # File 1: Buguianma
    wb1 = openpyxl.Workbook()
    ws1 = wb1.active
    ws1.title = "Spare Parts Mark"
    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 35
    ws1.column_dimensions["C"].width = 24
    ws1.column_dimensions["D"].width = 20
    ws1.column_dimensions["E"].width = 28

    brand_name = "XXXLutz"
    for ri, rd in enumerate([
        ["Delivery at:", "Central warehouse", brand_name, "Uffenheim", "Germany"],
        ["Supplier name:", "Deyuan Furniture", "Mark (Order No.): ", order_no, ""],
        ["Origin country:", "CHINA", "Container No./Tracking No.:", "", ""]
    ]):
        for ci, val in enumerate(rd):
            c = ws1.cell(row=ri + 1, column=ci + 1, value=val)
            c.font = h_fb if (ci >= 2 and ri in (0, 2)) else h_f
            c.alignment = ac
            c.border = tb
            if ri == 0:
                c.fill = yf

    ws1.merge_cells("C3:D3")
    ws1.cell(row=3, column=3).font = h_fb
    ws1.cell(row=3, column=3).alignment = ac
    ws1.cell(row=3, column=3).border = tb
    ws1.cell(row=3, column=4).border = tb

    for ci, val in enumerate(["Brand", "Description", "Photo", "Quantity", " Article No."]):
        c = ws1.cell(row=4, column=ci + 1, value=val)
        c.font = ch_fa if ci == 0 else ch_f
        c.alignment = ac
        c.border = tb

    cr = 5
    img1 = 0
    for key, group in groups.items():
        md = master_data.get(key)
        if not md:
            continue
        ns = set_counts[key]
        article = article_strs[key]
        for pi, part in enumerate(md["parts"]):
            r = cr + pi
            for ci in range(1, 6):
                ws1.cell(row=r, column=ci).border = tb
            ws1.cell(row=r, column=2, value=part["desc"]).font = d_f
            ws1.cell(row=r, column=2).alignment = al
            total = ns * part["per_set"]
            qt = "%dSETS\n%dX%d=%dPCS" % (ns, ns, part["per_set"], total)
            ws1.cell(row=r, column=4, value=qt).font = q_f
            ws1.cell(row=r, column=4).alignment = acw
            ws1.cell(row=r, column=5, value=article).font = a_f
            ws1.cell(row=r, column=5).alignment = alw
            for im in md["images"]:
                if im["from_row"] == 4 + pi:
                    try:
                        add_image(ws1, im["data"], r, "C", 24, 60)
                        img1 += 1
                    except:
                        pass
                    break
            ws1.row_dimensions[r].height = 60
        cr += len(md["parts"])

    out1 = os.path.join(out_dir, "补件唛头%s.xlsx" % order_no)
    wb1.save(out1)
    print("Saved: " + out1)
    print("  Images: " + str(img1))

    # Build image map
    ni_map = OrderedDict()
    for key, group in groups.items():
        md = master_data.get(key)
        if not md:
            continue
        for pi, part in enumerate(md["parts"]):
            norm = re.sub(r"^[A-Z]\s+[A-Za-z0-9]+\s+", "", part["desc"])
            norm = re.sub(r"^[A-Z]\s+", "", norm).strip()
            if not norm or len(norm) < 3:
                norm = part["desc"]
            if norm not in ni_map:
                for im in md["images"]:
                    if im["from_row"] == 4 + pi:
                        ni_map[norm] = im["data"]
                        break
    if "FEET" in ni_map and "Feet" not in ni_map:
        ni_map["Feet"] = ni_map["FEET"]
        del ni_map["FEET"]

    # File 2: SPL
    all_parts = []
    for key, group in groups.items():
        md = master_data.get(key)
        if not md:
            continue
        ns2 = set_counts[key]
        article2 = article_strs[key]
        for pi, part in enumerate(md["parts"]):
            total2 = ns2 * part["per_set"]
            norm = re.sub(r"^[A-Z]\s+[A-Za-z0-9]+\s+", "", part["desc"])
            norm = re.sub(r"^[A-Z]\s+", "", norm).strip()
            if not norm or len(norm) < 3:
                norm = part["desc"]
            img_d = None
            for im in md["images"]:
                if im["from_row"] == 4 + pi:
                    img_d = im["data"]
                    break
            all_parts.append({"norm": norm, "total": total2, "article": article2, "img": img_d})

    ft = 0
    fa = []
    fi = None
    fp = []
    for p in all_parts:
        if p["norm"].upper() == "FEET":
            ft += p["total"]
            fa.append(p["article"])
            if p["img"] and not fi:
                fi = p["img"]
        else:
            fp.append(p)
    if ft > 0:
        fp.append({"norm": "Feet", "total": ft,
                    "article": "\n".join(list(OrderedDict.fromkeys(fa))), "img": fi})
    all_parts = fp

    agg = OrderedDict()
    am = OrderedDict()
    ai = OrderedDict()
    for p in all_parts:
        if p["norm"] not in agg:
            agg[p["norm"]] = 0
            am[p["norm"]] = []
            ai[p["norm"]] = None
        agg[p["norm"]] += p["total"]
        am[p["norm"]].append(p["article"])
        if p["img"] and not ai[p["norm"]]:
            ai[p["norm"]] = p["img"]

    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "Spare Parts"
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 27.82
    ws2.column_dimensions["C"].width = 28
    ws2.column_dimensions["D"].width = 20
    ws2.column_dimensions["E"].width = 30
    ws2.column_dimensions["F"].width = 12

    for ri, rd in enumerate([
        ["Delivery at", "Centralwarehouse", brand_name, "", "Germany", ""],
        ["Supplier name:", "Deyuan Furniture", "Containerno.", "", "", ""],
        ["Delivery date:", "", "", "", "", ""]
    ]):
        for ci, val in enumerate(rd):
            c = ws2.cell(row=ri + 1, column=ci + 1, value=val)
            if ri == 0:
                cf = Font(name="Arial", size=16, color="FF000000", bold=True)
            elif ri == 1:
                if ci >= 3:
                    cf = Font(name="Arial", size=18, color="FF000000", bold=True)
                else:
                    cf = Font(name="Arial", size=16, color="FF000000", bold=True)
            else:
                cf = Font(name="Arial", size=16, color="FF000000", bold=True)
            c.font = cf
            c.alignment = ac if val else Alignment(horizontal="left", vertical="center")
            c.border = tb

    ws2.merge_cells("C1:D1")
    ws2.merge_cells("C2:D2")
    ws2.merge_cells("C3:D3")

    for ci, val in enumerate(["Description", "Material", "Photo", "Qty (pcs)",
                               "Used for Article No.", "Price "]):
        c = ws2.cell(row=4, column=ci + 1, value=val)
        if ci == 5:
            c.font = Font(name="Times New Roman", size=16, color="FF000000", bold=True)
        else:
            c.font = Font(name="Times New Roman", size=14, color="FF000000", bold=True)
        c.alignment = ac if ci < 4 else acw
        c.border = tb

    dr = 5
    img2 = 0
    for desc, q in agg.items():
        arts = list(OrderedDict.fromkeys(am[desc]))
        n_l = max(1, len("\n".join(arts).split("\n")))
        rh2 = max(70, n_l * 18)

        ws2.cell(row=dr, column=1, value=" " + desc).font = Font(
            name="SimSun", size=11, color="FF000000")
        ws2.cell(row=dr, column=1).alignment = ac
        ws2.cell(row=dr, column=1).border = tb

        ws2.cell(row=dr, column=2, value="Metal").font = Font(
            name="Times New Roman", size=14, color="FF000000")
        ws2.cell(row=dr, column=2).alignment = ac
        ws2.cell(row=dr, column=2).border = tb

        if ai.get(desc):
            try:
                add_image(ws2, ai[desc], dr, "C", 28, rh2)
                img2 += 1
            except:
                pass
        ws2.cell(row=dr, column=3).border = tb

        ws2.cell(row=dr, column=4, value=float(q)).font = Font(
            name="Arial", size=14, color="FF000000")
        ws2.cell(row=dr, column=4).alignment = ac
        ws2.cell(row=dr, column=4).border = tb

        ws2.cell(row=dr, column=5, value="\n".join(arts)).font = Font(
            name="Arial", size=14, color="FF000000")
        ws2.cell(row=dr, column=5).alignment = acw
        ws2.cell(row=dr, column=5).border = tb

        ws2.cell(row=dr, column=6, value=round(q * 0.001, 3)).font = Font(
            name="Arial", size=12, color="FF000000")
        ws2.cell(row=dr, column=6).alignment = acw
        ws2.cell(row=dr, column=6).border = tb

        ws2.row_dimensions[dr].height = rh2
        dr += 1

    ws2.cell(row=dr, column=1,
             value="We hereby declare: The Spare parts are free of charge. "
                   "The value is only for Customs clearance.").font = Font(
        name="SimSun", size=12, color="FF000000")
    ws2.cell(row=dr, column=1).alignment = ac
    ws2.cell(row=dr, column=1).border = tb
    ws2.row_dimensions[dr].height = 40

    spl_dir = os.path.join(out_dir, "请款资料")
    os.makedirs(spl_dir, exist_ok=True)
    out2 = os.path.join(spl_dir, "Spare Parts List-%s.xlsx" % order_no)
    wb2.save(out2)
    print("Saved: " + out2)
    print("  Images: " + str(img2))
    print("\n=== ALL DONE ===")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate LUTZ spare parts documents")
    parser.add_argument("--pi", required=True, help="Path to PI XLSX")
    parser.add_argument("--master", default=DEFAULT_MASTER, help="Path to master reference")
    parser.add_argument("--order-no", default=None, help="Order number")
    args = parser.parse_args()
    generate(args.pi, args.master, args.order_no)
